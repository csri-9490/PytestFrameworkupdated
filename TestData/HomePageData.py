import openpyxl
class HomePageData1:
    test_homepage_data=[{"Name":"csr1","email":"cs1@gamil.com","Password":"cs1234","Gender":"Female"},{"Name":"csr2","email":"cs2@gamil.com","Password":"cs1234","Gender":"Female"}]
    @staticmethod
    def getTestData(test_case_name): #self not required, beacuse above @staticmethod defined
        dict={}
        book=openpyxl.load_workbook("C:\\Users\\srika\\OneDrive\\Desktop\\Atmntestdta\\formtestdata.xlsx")
        sheet=book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[dict]

