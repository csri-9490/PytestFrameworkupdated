import openpyxl
book=openpyxl.load_workbook("C:\\Users\\srika\\OneDrive\\Desktop\\Atmntestdta\\exceldata.xlsx")
sheet=book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
# sheet.cell(row=2,column=2).value="csri"
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A2'].value)
# for i in range(1,sheet.max_row+1):
#     print(sheet.cell(row=i,column=i).value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="fname3":
     for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)
